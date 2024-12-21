import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re
import os

#previous
datasheet_name = 'data.xlsx'
if os.path.exists(datasheet_name):
    wb = load_workbook(datasheet_name)
    ws = wb.active
else:    
#create Excel book
    wb = Workbook()
    ws = wb.active
    ws.append(["Name","Age","Mail","Phone","Direction"])

def save_data():
    name = entry_name.get()
    age = entry_age.get()
    mail = entry_mail.get()
    phone = entry_phone.get()
    direction = entry_direction.get()
    
    #validate empty fields
    if not name or not age or not mail or not phone or not direction:
        messagebox.showwarning("Alert", "Empty field")
        return
    
    #validate data type
    try:
        age = int(age)
        phone = int(phone)
    except ValueError:
        messagebox.showwarning("Alert", "Age and Phone would be numbers")
        return
    
    #validate mail
    if not re.match(r"[^@]+@[^@]+\.[^@]+", mail):
        messagebox.showwarning("Alert", "Invalid mail")
        return
    
    ws.append([name, age, mail, phone, direction])
    wb.save(datasheet_name)
    messagebox.showinfo("Information", "Data stored successfully")
    
    entry_name.delete(0, tk.END)
    entry_age.delete(0, tk.END)
    entry_mail.delete(0, tk.END)
    entry_phone.delete(0, tk.END)
    entry_direction.delete(0, tk.END)

root = tk.Tk()
root.title("Form")
root.configure(bg='#4B6587')
label_style = {"bg" : '#4B6587', "fg" : 'white'}
entry_style = {"bg" : '#D3D3D3', "fg" : 'black'}

label_name = tk.Label(root, text="Name", **label_style)
label_name.grid(row = 0, column = 0, padx = 10, pady = 5)
entry_name = tk.Entry(root, **entry_style)
entry_name.grid(row = 0, column = 1, padx = 10, pady = 5)

label_age = tk.Label(root, text="Age", **label_style)
label_age.grid(row = 1, column = 0, padx = 10, pady = 5)
entry_age = tk.Entry(root, **entry_style)
entry_age.grid(row = 1, column = 1, padx = 10, pady = 5)

label_mail = tk.Label(root, text="Mail", **label_style)
label_mail.grid(row = 2, column = 0, padx = 10, pady = 5)
entry_mail = tk.Entry(root, **entry_style)
entry_mail.grid(row = 2, column = 1, padx = 10, pady = 5)

label_phone = tk.Label(root, text="Phone", **label_style)
label_phone.grid(row = 3, column = 0, padx = 10, pady = 5)
entry_phone = tk.Entry(root, **entry_style)
entry_phone.grid(row = 3, column = 1, padx = 10, pady = 5)

label_direction = tk.Label(root, text="Direction", **label_style)
label_direction.grid(row = 4, column = 0, padx = 10, pady = 5)
entry_direction = tk.Entry(root, **entry_style)
entry_direction.grid(row = 4, column = 1, padx = 10, pady = 5)

btn_save = tk.Button(root, text="Save", command = save_data, bg ='#6D8299', fg = 'white')
btn_save.grid(row = 5, column = 0, columnspan = 2, padx = 10, pady = 10)

root.mainloop()